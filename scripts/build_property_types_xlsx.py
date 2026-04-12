"""
Build a clean Excel file of Crexi-style CRE property types and subtypes.
Outputs: crexi-property-types.xlsx at the repo root.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# (main_type, subtype, slug, crexi_verified, notes)
ROWS = [
    # Retail
    ("Retail", "Retail (All)", "retail", True, "Top-level Crexi category"),
    ("Retail", "Shopping Center", "shopping-centers", True, "Generic shopping center"),
    ("Retail", "Neighborhood Center", "neighborhood-center", False, "Under shopping centers"),
    ("Retail", "Community Center", "community-center", False, "Under shopping centers"),
    ("Retail", "Power Center", "power-center", False, "Big-box anchored"),
    ("Retail", "Regional / Super-Regional Mall", "regional-mall", False, ""),
    ("Retail", "Lifestyle Center", "lifestyle-center", False, ""),
    ("Retail", "Outlet Center", "outlet-center", False, ""),
    ("Retail", "Strip Center / Strip Mall", "strip-center", False, ""),
    ("Retail", "Freestanding / Single-Tenant", "freestanding-retail", False, "Often NNN"),
    ("Retail", "Restaurant / QSR", "restaurant", False, "Quick-service restaurant"),
    ("Retail", "Bank", "bank", False, ""),
    ("Retail", "Auto Dealership", "auto-dealership", False, ""),
    ("Retail", "Convenience Store / C-Store", "convenience-store", False, ""),
    ("Retail", "Pharmacy / Drugstore", "pharmacy", False, ""),
    ("Retail", "Big Box / Anchor", "big-box", False, ""),
    ("Retail", "Pad Site / Outparcel", "pad-site", False, ""),
    ("Retail", "High Street / Storefront", "storefront", False, ""),

    # Office
    ("Office", "Office (All)", "office", True, "Top-level Crexi category"),
    ("Office", "Traditional Office Class A", "office-class-a", False, ""),
    ("Office", "Traditional Office Class B", "office-class-b", False, ""),
    ("Office", "Traditional Office Class C", "office-class-c", False, ""),
    ("Office", "Medical Office Building (MOB)", "medical-office", False, ""),
    ("Office", "Creative Office / Loft", "creative-office", False, ""),
    ("Office", "Executive Suites / Coworking", "coworking", False, ""),
    ("Office", "Professional / Dental", "professional-office", False, ""),
    ("Office", "Office Condo", "office-condo", False, ""),
    ("Office", "Government Office", "government-office", False, ""),
    ("Office", "Office Campus", "office-campus", False, ""),

    # Industrial
    ("Industrial", "Industrial (All)", "industrial", True, "Top-level Crexi category"),
    ("Industrial", "Warehouse", "warehouses", True, "Verified Crexi slug"),
    ("Industrial", "Distribution Center", "distribution-center", False, ""),
    ("Industrial", "Manufacturing", "manufacturing", False, ""),
    ("Industrial", "Flex / R&D", "industrial-flex", True, "Verified Crexi slug"),
    ("Industrial", "Light Industrial", "light-industrial", False, ""),
    ("Industrial", "Heavy Industrial", "heavy-industrial", False, ""),
    ("Industrial", "Refrigerated / Cold Storage", "refrigerated-and-cold-storage", True, "Verified Crexi slug"),
    ("Industrial", "Truck Terminal / Transportation", "truck-terminal", False, ""),
    ("Industrial", "Data Center", "data-center", False, ""),
    ("Industrial", "Industrial Outdoor Storage (IOS)", "ios", False, "Growing asset class"),

    # Multifamily
    ("Multifamily", "Multifamily (All)", "multifamily", True, "Top-level Crexi category"),
    ("Multifamily", "Garden-Style (1-3 stories)", "garden-apartments", False, ""),
    ("Multifamily", "Mid-Rise (4-9 stories)", "mid-rise-apartments", False, ""),
    ("Multifamily", "High-Rise (10+ stories)", "high-rise-apartments", False, ""),
    ("Multifamily", "Low-Rise / Walk-Up", "walk-up-apartments", False, ""),
    ("Multifamily", "Duplex / Triplex / Fourplex", "small-multifamily", False, "2-4 units"),
    ("Multifamily", "Apartment Complex (5+ units)", "apartment-complex", False, ""),
    ("Multifamily", "Student Housing", "student-housing", True, "Verified Crexi slug"),
    ("Multifamily", "Affordable / LIHTC Housing", "affordable-housing", False, ""),
    ("Multifamily", "Military Housing", "military-housing", False, ""),
    ("Multifamily", "Condominium Building", "condo-building", False, ""),
    ("Multifamily", "Co-op Building", "coop-building", False, ""),
    ("Multifamily", "Mixed-Use Multifamily", "mixed-use-multifamily", False, "Apartments over retail"),

    # Hospitality
    ("Hospitality", "Hospitality (All)", "hospitality", True, "Top-level Crexi category"),
    ("Hospitality", "Full-Service Hotel", "full-service-hotel", False, ""),
    ("Hospitality", "Limited-Service Hotel", "limited-service-hotel", False, ""),
    ("Hospitality", "Select-Service Hotel", "select-service-hotel", False, ""),
    ("Hospitality", "Boutique Hotel", "boutique-hotel", False, ""),
    ("Hospitality", "Extended-Stay Hotel", "extended-stay-hotel", False, ""),
    ("Hospitality", "Resort", "resort", False, ""),
    ("Hospitality", "Motel", "motel", False, ""),
    ("Hospitality", "Bed & Breakfast / Inn", "bed-and-breakfast", False, ""),
    ("Hospitality", "Casino / Gaming", "casino", False, ""),
    ("Hospitality", "Hostel", "hostel", False, ""),

    # Land
    ("Land", "Land (All)", "land", True, "Top-level Crexi category"),
    ("Land", "Commercial Land", "commercial-land", True, "Verified Crexi slug"),
    ("Land", "Residential Land / Development Parcel", "residential-land", False, ""),
    ("Land", "Industrial Land", "industrial-land", False, ""),
    ("Land", "Retail Pad / Outparcel", "retail-pad", False, ""),
    ("Land", "Mixed-Use Development Land", "mixed-use-land", False, ""),
    ("Land", "Agricultural Land", "agricultural-land", True, "Verified Crexi slug"),
    ("Land", "Ranch / Farm", "ranch-farm", False, ""),
    ("Land", "Timberland", "timberland", False, ""),
    ("Land", "Hunting / Recreational Land", "recreational-land", False, ""),
    ("Land", "Waterfront / Coastal Land", "waterfront-land", False, ""),
    ("Land", "Entitled Land", "entitled-land", False, ""),
    ("Land", "Raw Land / Unentitled", "raw-land", False, ""),

    # Mixed Use
    ("Mixed Use", "Mixed Use (All)", "mixed-use", True, "Top-level Crexi category"),
    ("Mixed Use", "Retail + Office", "retail-office-mixed", False, ""),
    ("Mixed Use", "Retail + Residential", "retail-residential-mixed", False, ""),
    ("Mixed Use", "Office + Residential", "office-residential-mixed", False, ""),
    ("Mixed Use", "Live / Work", "live-work", False, ""),
    ("Mixed Use", "Vertical Mixed-Use", "vertical-mixed-use", False, ""),

    # Special Purpose
    ("Special Purpose", "Special Purpose (All)", "special-purpose", True, "Top-level Crexi category"),
    ("Special Purpose", "Car Wash", "car-washes", True, "Verified Crexi slug"),
    ("Special Purpose", "Gas Station", "gas-station", True, "Verified Crexi slug"),
    ("Special Purpose", "Church / Religious Facility", "religious-facility", False, ""),
    ("Special Purpose", "School / Daycare / Childcare", "school-daycare", False, ""),
    ("Special Purpose", "Auto Repair / Body Shop / Lube", "auto-service", False, ""),
    ("Special Purpose", "Funeral Home / Cemetery", "funeral-home", False, ""),
    ("Special Purpose", "Marina / Boat Storage", "marina", False, ""),
    ("Special Purpose", "Parking Garage / Lot", "parking-garage", False, ""),
    ("Special Purpose", "Fitness / Health Club / Gym", "fitness-center", False, ""),
    ("Special Purpose", "Movie Theater / Cinema", "movie-theater", False, ""),
    ("Special Purpose", "Golf Course / Country Club", "golf-course", False, ""),
    ("Special Purpose", "Bowling Alley", "bowling-alley", False, ""),
    ("Special Purpose", "Amusement / Entertainment", "amusement", False, ""),
    ("Special Purpose", "Sports Facility / Arena", "sports-facility", False, ""),
    ("Special Purpose", "RV Park / Campground", "rv-park", False, "Sometimes under Mobile Home"),
    ("Special Purpose", "Winery / Vineyard", "winery", False, ""),
    ("Special Purpose", "Distillery / Brewery", "distillery-brewery", False, ""),
    ("Special Purpose", "Nursery / Garden Center", "nursery", False, ""),
    ("Special Purpose", "Cannabis-Compliant Facility", "cannabis-facility", False, ""),
    ("Special Purpose", "Post Office", "post-office", False, ""),
    ("Special Purpose", "Veterinary / Animal Hospital / Kennel", "veterinary", False, ""),

    # Self Storage
    ("Self Storage", "Self Storage (All)", "self-storage", True, "Top-level Crexi category"),
    ("Self Storage", "Drive-Up Self Storage", "drive-up-storage", False, ""),
    ("Self Storage", "Climate-Controlled Self Storage", "climate-storage", False, ""),
    ("Self Storage", "Boat & RV Storage", "boat-rv-storage", False, ""),
    ("Self Storage", "Mixed Self Storage Facility", "mixed-storage", False, ""),

    # Manufactured Housing / Mobile Home Park
    ("Manufactured Housing", "Mobile Home Park (All)", "mobile-home-park", True, "Top-level Crexi category"),
    ("Manufactured Housing", "All-Age Park", "all-age-park", False, ""),
    ("Manufactured Housing", "Age-Restricted 55+ Park", "age-restricted-park", False, ""),
    ("Manufactured Housing", "Family Park", "family-park", False, ""),
    ("Manufactured Housing", "RV Park", "rv-park-mh", False, "Also under Special Purpose"),

    # Senior Living / Healthcare
    ("Senior Living / Healthcare", "Senior Living (All)", "senior-living", True, "Top-level Crexi category"),
    ("Senior Living / Healthcare", "Independent Living", "independent-living", False, ""),
    ("Senior Living / Healthcare", "Assisted Living Facility (ALF)", "assisted-living", False, ""),
    ("Senior Living / Healthcare", "Memory Care", "memory-care", False, ""),
    ("Senior Living / Healthcare", "Skilled Nursing Facility (SNF)", "skilled-nursing", False, ""),
    ("Senior Living / Healthcare", "Continuing Care Retirement Community", "ccrc", False, ""),
    ("Senior Living / Healthcare", "Hospital", "hospital", False, ""),
    ("Senior Living / Healthcare", "Surgery Center", "surgery-center", False, "Ambulatory surgical center"),
    ("Senior Living / Healthcare", "Urgent Care", "urgent-care", False, ""),
    ("Senior Living / Healthcare", "Rehab / LTAC", "rehab-ltac", False, "Long-term acute care"),

    # Business for Sale
    ("Business for Sale", "Business for Sale (All)", "business-for-sale", True, "Top-level Crexi category"),
    ("Business for Sale", "Car Wash Business", "car-wash-business", False, "Operating car wash"),
    ("Business for Sale", "Restaurant / Bar", "restaurant-business", False, ""),
    ("Business for Sale", "Franchise Resale", "franchise-resale", False, ""),
    ("Business for Sale", "Liquor Store / Package Store", "liquor-store", False, ""),
    ("Business for Sale", "Laundromat", "laundromat", False, ""),
    ("Business for Sale", "Gym / Fitness Business", "gym-business", False, ""),
    ("Business for Sale", "C-Store / Gas Business", "c-store-business", False, ""),
    ("Business for Sale", "Hotel / Motel Business", "hotel-business", False, ""),
    ("Business for Sale", "Auto Service Business", "auto-service-business", False, ""),

    # Note / Loan
    ("Note / Loan", "Note / Loan (All)", "note-loan", True, "Top-level Crexi category"),
    ("Note / Loan", "Performing Note", "performing-note", False, ""),
    ("Note / Loan", "Non-Performing Note (NPL)", "non-performing-note", False, ""),
    ("Note / Loan", "Reperforming Note", "reperforming-note", False, ""),
    ("Note / Loan", "Commercial Mortgage", "commercial-mortgage", False, ""),
    ("Note / Loan", "Residential Mortgage Pool", "residential-mortgage-pool", False, ""),
]


def build():
    wb = Workbook()

    # === Sheet 1: Full taxonomy ===
    ws = wb.active
    ws.title = "Property Types"

    headers = [
        "Main Type",
        "Subtype",
        "Slug",
        "Crexi Verified",
        "Crexi URL",
        "Notes",
    ]
    ws.append(headers)

    # Styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F3864")
    header_align = Alignment(horizontal="left", vertical="center")
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Data rows
    for row in ROWS:
        main_type, subtype, slug, verified, notes = row
        crexi_url = (
            f"https://www.crexi.com/properties/{slug.replace('-', '_').title().replace('_', '-')}"
            if verified
            else ""
        )
        # For verified ones we know the exact slugs — hardcode the known ones
        # to match what Crexi actually uses.
        known_slugs = {
            "retail": "Retail",
            "office": "Office",
            "industrial": "Industrial",
            "multifamily": "Multifamily",
            "hospitality": "Hospitality",
            "land": "Land",
            "mixed-use": "Mixed-Use",
            "special-purpose": "Special-Purpose",
            "self-storage": "Self-Storage",
            "mobile-home-park": "Mobile-Home-Park",
            "senior-living": "Senior-Living",
            "business-for-sale": "Business-For-Sale",
            "note-loan": "Note-Loan",
            "shopping-centers": "Shopping-Centers",
            "warehouses": "Warehouses",
            "industrial-flex": "Industrial-Flex",
            "refrigerated-and-cold-storage": "Refrigerated-and-Cold-Storage",
            "student-housing": "Student-Housing",
            "commercial-land": "Commercial-Land",
            "agricultural-land": "Agricultural-Land",
            "car-washes": "Car-Washes",
            "gas-station": "Gas-Station",
        }
        if slug in known_slugs and verified:
            crexi_url = f"https://www.crexi.com/properties/{known_slugs[slug]}"
        elif not verified:
            crexi_url = ""

        ws.append([
            main_type,
            subtype,
            slug,
            "Yes" if verified else "No",
            crexi_url,
            notes,
        ])

    # Apply body styling + alt row banding
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="F2F2F2")

    # Highlight "Crexi Verified = Yes" rows
    green_fill = PatternFill("solid", start_color="E2EFDA")
    verified_col = 4  # "Crexi Verified"
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=verified_col).value == "Yes":
            ws.cell(row=row_idx, column=verified_col).fill = green_fill
            ws.cell(row=row_idx, column=verified_col).font = Font(
                name="Arial", size=10, bold=True, color="375623"
            )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add autofilter
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    col_widths = {"A": 26, "B": 42, "C": 34, "D": 15, "E": 58, "F": 40}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 24

    # === Sheet 2: Summary by main type ===
    summary = wb.create_sheet("Summary")
    summary.append(["Main Type", "Subtype Count", "Crexi Verified Count"])
    for col_idx in range(1, 4):
        cell = summary.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    main_types_in_order = []
    for row in ROWS:
        if row[0] not in main_types_in_order:
            main_types_in_order.append(row[0])

    for row_idx, mt in enumerate(main_types_in_order, start=2):
        summary.cell(row=row_idx, column=1, value=mt)
        # Use Excel formulas for counts — stay dynamic
        summary.cell(
            row=row_idx,
            column=2,
            value=f'=COUNTIF(\'Property Types\'!A:A,A{row_idx})',
        )
        summary.cell(
            row=row_idx,
            column=3,
            value=f'=COUNTIFS(\'Property Types\'!A:A,A{row_idx},\'Property Types\'!D:D,"Yes")',
        )

    # Totals row
    total_row = len(main_types_in_order) + 2
    summary.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name="Arial", size=10, bold=True
    )
    summary.cell(
        row=total_row,
        column=2,
        value=f"=SUM(B2:B{total_row - 1})",
    ).font = Font(name="Arial", size=10, bold=True)
    summary.cell(
        row=total_row,
        column=3,
        value=f"=SUM(C2:C{total_row - 1})",
    ).font = Font(name="Arial", size=10, bold=True)

    # Style summary body
    for row_idx in range(2, total_row + 1):
        for col_idx in range(1, 4):
            cell = summary.cell(row=row_idx, column=col_idx)
            if not cell.font or cell.font.name != "Arial":
                cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_idx == total_row:
                cell.fill = PatternFill("solid", start_color="FFE699")

    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 22
    summary.freeze_panes = "A2"

    # === Sheet 3: Legend / notes ===
    legend = wb.create_sheet("Legend")
    legend_rows = [
        ("Column", "Description"),
        ("Main Type", "Crexi's top-level property category (Retail, Office, etc.)"),
        ("Subtype", "Specific sub-classification within the main type"),
        ("Slug", "URL-safe identifier suitable for code / routing"),
        ("Crexi Verified", "Yes = the slug was confirmed in a real Crexi URL during research"),
        ("Crexi URL", "Direct link to Crexi category page (verified slugs only)"),
        ("Notes", "Source commentary or cross-reference to other categories"),
        ("", ""),
        ("Legend", ""),
        ("Yes (green)", "Slug verified against a live Crexi URL"),
        ("No", "Standard industry subtype — Crexi may categorize these under the parent"),
    ]
    for row in legend_rows:
        legend.append(row)

    # Style legend
    for col_idx in range(1, 3):
        cell = legend.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    for row_idx in range(2, legend.max_row + 1):
        for col_idx in range(1, 3):
            cell = legend.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    legend.cell(row=9, column=1).font = Font(name="Arial", size=10, bold=True)
    legend.column_dimensions["A"].width = 24
    legend.column_dimensions["B"].width = 70

    out_path = "crexi-property-types.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path} with {len(ROWS)} rows across {len(main_types_in_order)} main types")


if __name__ == "__main__":
    build()
